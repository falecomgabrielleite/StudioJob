import os
import openpyxl

diretorio_principal = 'X:\Eloisa - Estudio\BRANDILI - PREMISSAS OI25\Brandili 3744X5616'
arquivo_excel = 'renomear-dia2.xlsx'

# Inicialize um objeto Workbook
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Adicione um cabeçalho à primeira linha
worksheet['A1'] = 'Caminho '
worksheet['B1'] = 'Nome do arquivo'
worksheet['C1'] = 'Caminho Completo'


# Inicialize uma variável para rastrear a linha atual
linha_atual = 2

# Percorra todos os arquivos e subdiretórios no diretório principal
#for pasta_atual, _, arquivos in os.walk(diretorio_principal):
#    for arquivo in arquivos:
#        caminho_completo = os.path.join(pasta_atual, arquivo)
#        # Adicione o caminho completo do arquivo à planilha
#        worksheet.cell(row=linha_atual, column=1, value=os.path.dirname(caminho_completo))
#        # Adicione o nome do arquivo à coluna B
#        worksheet.cell(row=linha_atual, column=2, value=os.path.basename(arquivo))
#        # Adicione o diretório à coluna C
#        worksheet.cell(row=linha_atual, column=3, value=caminho_completo)
#        linha_atual += 1

# Percorra todos os arquivos e subdiretórios no diretório principal
try:
    for pasta_atual, _, arquivos in os.walk(diretorio_principal):
        for arquivo in arquivos:
            caminho_completo = os.path.join(pasta_atual, arquivo)
            # Adicione o caminho completo do arquivo à planilha
            worksheet.cell(row=linha_atual, column=1, value=pasta_atual)
            # Adicione o nome do arquivo à coluna B
            worksheet.cell(row=linha_atual, column=2, value=arquivo)
            # Adicione o caminho completo do arquivo à coluna C
            worksheet.cell(row=linha_atual, column=3, value=caminho_completo)
            linha_atual += 1
except Exception as e:
    print(f"Ocorreu um erro: {e}")

# Salve o arquivo Excel
workbook.save(arquivo_excel)

print(f'A lista de arquivos foi salva em "{arquivo_excel}" com sucesso!')
